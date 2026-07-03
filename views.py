import csv
import io
from datetime import datetime, timedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.db.models import Count, Sum, Q
from django.utils import timezone

from .models import Candidate, Vote
from .forms import AdminLoginForm, CandidateForm


def get_client_info(request):
    """Extract browser, OS, device info from request."""
    ua = request.META.get('HTTP_USER_AGENT', '')

    browser = 'Unknown'
    if 'Firefox' in ua:
        browser = 'Firefox'
    elif 'Edg' in ua:
        browser = 'Edge'
    elif 'Chrome' in ua:
        browser = 'Chrome'
    elif 'Safari' in ua:
        browser = 'Safari'
    elif 'Opera' in ua or 'OPR' in ua:
        browser = 'Opera'

    os_name = 'Unknown'
    if 'Windows' in ua:
        os_name = 'Windows'
    elif 'Mac OS' in ua:
        os_name = 'macOS'
    elif 'Linux' in ua:
        os_name = 'Linux'
    elif 'Android' in ua:
        os_name = 'Android'
    elif 'iPhone' in ua or 'iPad' in ua:
        os_name = 'iOS'

    device = 'Desktop'
    if 'Mobile' in ua or 'Android' in ua:
        device = 'Mobile'
    elif 'iPad' in ua or 'Tablet' in ua:
        device = 'Tablet'

    return browser, os_name, device


# ─── Admin Login/Logout ────────────────────────────────────────────

class AdminLoginView(View):
    def get(self, request):
        if request.user.is_authenticated:
            return redirect('voting:admin_dashboard')
        form = AdminLoginForm()
        return render(request, 'voting/admin_login.html', {'form': form})

    def post(self, request):
        form = AdminLoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('voting:admin_dashboard')
        return render(request, 'voting/admin_login.html', {'form': form})


@login_required
def admin_logout_view(request):
    logout(request)
    messages.success(request, 'You have been logged out.')
    return redirect('voting:admin_login')


# ─── Admin Dashboard ───────────────────────────────────────────────

class AdminDashboardView(LoginRequiredMixin, View):
    login_url = 'voting:admin_login'

    def get(self, request):
        total_candidates = Candidate.objects.count()
        total_votes = Vote.objects.count()
        today = timezone.now().date()
        today_votes = Vote.objects.filter(vote_time__date=today).count()
        remaining = max(0, 1000 - total_votes)

        # Winner
        winner_data = (
            Vote.objects.values('candidate__name')
            .annotate(total=Count('id'))
            .order_by('-total')
            .first()
        )
        winner_name = winner_data['candidate__name'] if winner_data else 'N/A'
        winner_votes = winner_data['total'] if winner_data else 0

        # Latest vote
        latest_vote = Vote.objects.select_related('candidate').first()

        context = {
            'total_candidates': total_candidates,
            'total_votes': total_votes,
            'today_votes': today_votes,
            'remaining_votes': remaining,
            'winner_name': winner_name,
            'winner_votes': winner_votes,
            'latest_vote': latest_vote,
        }
        return render(request, 'voting/admin_dashboard.html', context)


# ─── Candidate Management ──────────────────────────────────────────

class CandidateListView(LoginRequiredMixin, View):
    login_url = 'voting:admin_login'

    def get(self, request):
        candidates = Candidate.objects.all()
        search = request.GET.get('search', '').strip()
        filter_by = request.GET.get('filter', '')

        if search:
            candidates = candidates.filter(
                Q(name__icontains=search) |
                Q(party__icontains=search) |
                Q(position__icontains=search)
            )

        if filter_by == 'most_votes':
            candidates = candidates.annotate(vote_count=Count('votes')).order_by('-vote_count')
        elif filter_by == 'least_votes':
            candidates = candidates.annotate(vote_count=Count('votes')).order_by('vote_count')
        elif filter_by == 'newest':
            candidates = candidates.order_by('-created_at')
        elif filter_by == 'oldest':
            candidates = candidates.order_by('created_at')

        context = {
            'candidates': candidates,
            'search': search,
            'filter_by': filter_by,
        }
        return render(request, 'voting/candidate_list.html', context)


class CandidateCreateView(LoginRequiredMixin, View):
    login_url = 'voting:admin_login'

    def get(self, request):
        form = CandidateForm()
        return render(request, 'voting/candidate_form.html', {'form': form, 'action': 'Add'})

    def post(self, request):
        form = CandidateForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Candidate added successfully!')
            return redirect('voting:candidate_list')
        return render(request, 'voting/candidate_form.html', {'form': form, 'action': 'Add'})


class CandidateUpdateView(LoginRequiredMixin, View):
    login_url = 'voting:admin_login'

    def get(self, request, pk):
        candidate = get_object_or_404(Candidate, pk=pk)
        form = CandidateForm(instance=candidate)
        return render(request, 'voting/candidate_form.html', {'form': form, 'candidate': candidate, 'action': 'Edit'})

    def post(self, request, pk):
        candidate = get_object_or_404(Candidate, pk=pk)
        form = CandidateForm(request.POST, request.FILES, instance=candidate)
        if form.is_valid():
            form.save()
            messages.success(request, 'Candidate updated successfully!')
            return redirect('voting:candidate_list')
        return render(request, 'voting/candidate_form.html', {'form': form, 'candidate': candidate, 'action': 'Edit'})


class CandidateDeleteView(LoginRequiredMixin, View):
    login_url = 'voting:admin_login'

    def post(self, request, pk):
        candidate = get_object_or_404(Candidate, pk=pk)
        candidate.delete()
        messages.success(request, 'Candidate deleted successfully!')
        return redirect('voting:candidate_list')


class CandidateDetailView(LoginRequiredMixin, View):
    login_url = 'voting:admin_login'

    def get(self, request, pk):
        candidate = get_object_or_404(Candidate, pk=pk)
        votes = candidate.votes.all()[:50]
        total_votes = candidate.total_votes
        return render(request, 'voting/candidate_detail.html', {
            'candidate': candidate,
            'votes': votes,
            'total_votes': total_votes,
        })


# ─── Public Voting Page ────────────────────────────────────────────

class VotingPageView(View):
    def get(self, request):
        candidates = Candidate.objects.filter(status='active')
        return render(request, 'voting/voting_page.html', {'candidates': candidates})


@csrf_exempt
@require_POST
def cast_vote(request):
    """Handle vote submission via AJAX."""
    voter_id = request.POST.get('voter_id', '').strip()
    mobile_number = request.POST.get('mobile_number', '').strip()
    candidate_id = request.POST.get('candidate_id')
    voter_id_type = request.POST.get('voter_id_type', 'student_id')

    if not voter_id or not candidate_id:
        return JsonResponse({'success': False, 'message': 'All fields are required.'}, status=400)

    # Use voter_id as unique identifier
    unique_voter_id = voter_id

    # Check duplicate
    if Vote.objects.filter(voter_id=unique_voter_id).exists():
        return JsonResponse({'success': False, 'message': 'You have already voted.'}, status=400)

    candidate = get_object_or_404(Candidate, pk=candidate_id, status='active')
    browser, os_name, device = get_client_info(request)
    ip_address = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR', '')).split(',')[0].strip()

    Vote.objects.create(
        candidate=candidate,
        voter_id=unique_voter_id,
        mobile_number=mobile_number,
        ip_address=ip_address,
        browser=browser,
        operating_system=os_name,
        device=device,
    )

    return JsonResponse({
        'success': True,
        'message': 'Thank you for voting!',
        'candidate_name': candidate.name,
    })


# ─── Live Vote Counter (AJAX) ─────────────────────────────────────

@login_required
def live_vote_counter(request):
    """Return live vote data for AJAX polling."""
    candidates = Candidate.objects.filter(status='active')
    total_votes = Vote.objects.count()

    data = []
    for c in candidates:
        count = c.total_votes
        percentage = round((count / total_votes * 100), 1) if total_votes > 0 else 0
        data.append({
            'id': c.id,
            'name': c.name,
            'party': c.party or '',
            'votes': count,
            'percentage': percentage,
            'photo_url': c.photo.url if c.photo else '',
        })

    return JsonResponse({'candidates': data, 'total_votes': total_votes})


# ─── Result Page ───────────────────────────────────────────────────

class ResultPageView(View):
    def get(self, request):
        candidates = Candidate.objects.filter(status='active')
        total_votes = Vote.objects.count()

        winner = None
        winner_count = 0
        candidate_data = []

        for c in candidates:
            count = c.total_votes
            percentage = round((count / total_votes * 100), 1) if total_votes > 0 else 0
            candidate_data.append({
                'candidate': c,
                'votes': count,
                'percentage': percentage,
            })
            if count > winner_count:
                winner_count = count
                winner = c

        # Sort by votes descending
        candidate_data.sort(key=lambda x: x['votes'], reverse=True)

        context = {
            'winner': winner,
            'winner_votes': winner_count,
            'winner_percentage': round((winner_count / total_votes * 100), 1) if total_votes > 0 and winner else 0,
            'candidate_data': candidate_data,
            'total_votes': total_votes,
        }
        return render(request, 'voting/result_page.html', context)


# ─── Charts Data (AJAX) ───────────────────────────────────────────

@login_required
def charts_data(request):
    """Return chart data as JSON."""
    candidates = Candidate.objects.filter(status='active')
    total_votes = Vote.objects.count()

    labels = []
    votes_data = []
    colors = [
        '#6366f1', '#8b5cf6', '#ec4899', '#f43f5e', '#f97316',
        '#eab308', '#22c55e', '#14b8a6', '#06b6d4', '#3b82f6',
    ]

    for c in candidates:
        labels.append(c.name)
        votes_data.append(c.total_votes)

    return JsonResponse({
        'labels': labels,
        'votes': votes_data,
        'colors': colors[:len(labels)],
        'total_votes': total_votes,
    })


# ─── Reports ───────────────────────────────────────────────────────

@login_required
def report_pdf(request):
    """Generate PDF-like HTML report (using reportlab if available, else HTML)."""
    try:
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Voting Report", styles['Title']))
        elements.append(Spacer(1, 0.5 * inch))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 0.3 * inch))

        data = [['Candidate', 'Party', 'Position', 'Votes', 'Percentage']]
        total_votes = Vote.objects.count()
        for c in Candidate.objects.all():
            count = c.total_votes
            pct = f"{round(count / total_votes * 100, 1)}%" if total_votes > 0 else "0%"
            data.append([c.name, c.party or '-', c.position, str(count), pct])

        table = Table(data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#6366f1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), rl_colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 1, rl_colors.HexColor('#e2e8f0')),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="voting_report_{datetime.now().strftime("%Y%m%d")}.pdf"'
        return response

    except ImportError:
        # Fallback to HTML report
        return _generate_html_report(request)


@login_required
def report_excel(request):
    """Generate Excel report using openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Voting Report"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")

        headers = ['Candidate', 'Party', 'Position', 'Votes', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        total_votes = Vote.objects.count()
        for row, c in enumerate(Candidate.objects.all(), 2):
            count = c.total_votes
            pct = round(count / total_votes * 100, 1) if total_votes > 0 else 0
            ws.cell(row=row, column=1, value=c.name)
            ws.cell(row=row, column=2, value=c.party or '-')
            ws.cell(row=row, column=3, value=c.position)
            ws.cell(row=row, column=4, value=count)
            ws.cell(row=row, column=5, value=f"{pct}%")

        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="voting_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        return response

    except ImportError:
        return HttpResponse("Excel export requires openpyxl. pip install openpyxl", status=500)


@login_required
def report_csv(request):
    """Generate CSV report."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="voting_report_{datetime.now().strftime("%Y%m%d")}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Candidate', 'Party', 'Position', 'Votes', 'Percentage', 'Vote Time'])

    total_votes = Vote.objects.count()
    for c in Candidate.objects.all():
        count = c.total_votes
        pct = round(count / total_votes * 100, 1) if total_votes > 0 else 0
        last_vote = c.votes.first()
        vote_time = last_vote.vote_time.strftime('%Y-%m-%d %H:%M') if last_vote else '-'
        writer.writerow([c.name, c.party or '-', c.position, count, f"{pct}%", vote_time])

    return response


def _generate_html_report(request):
    """Fallback HTML report when reportlab is not installed."""
    candidates = Candidate.objects.all()
    total_votes = Vote.objects.count()

    html = f"""
    <html><head><title>Voting Report</title>
    <style>body{{font-family:Arial}}table{{border-collapse:collapse;width:100%}}th,td{{border:1px solid #ddd;padding:8px;text-align:center}}th{{background:#6366f1;color:white}}</style>
    </head><body>
    <h1>Voting Report</h1>
    <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
    <table><tr><th>Candidate</th><th>Party</th><th>Position</th><th>Votes</th><th>Percentage</th></tr>
    """
    for c in candidates:
        count = c.total_votes
        pct = round(count / total_votes * 100, 1) if total_votes > 0 else 0
        html += f"<tr><td>{c.name}</td><td>{c.party or '-'}</td><td>{c.position}</td><td>{count}</td><td>{pct}%</td></tr>"
    html += "</table></body></html>"

    response = HttpResponse(html, content_type='text/html')
    response['Content-Disposition'] = f'attachment; filename="voting_report_{datetime.now().strftime("%Y%m%d")}.html"'
    return response


# ─── Vote Status Check (AJAX) ─────────────────────────────────────

@csrf_exempt
def check_vote_status(request):
    """Check if a voter has already voted."""
    voter_id = request.POST.get('voter_id', '').strip()
    if not voter_id:
        return JsonResponse({'has_voted': False})

    has_voted = Vote.objects.filter(voter_id=voter_id).exists()
    voted_for = None
    if has_voted:
        vote = Vote.objects.filter(voter_id=voter_id).first()
        voted_for = vote.candidate.name

    return JsonResponse({
        'has_voted': has_voted,
        'voted_for': voted_for,
    })

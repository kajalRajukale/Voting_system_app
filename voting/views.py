import csv
import io
from datetime import datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.db.models import Count, F, Max
from django.db.models.functions import TruncDate

from .models import Candidate, Vote, MachineState
from .forms import AdminLoginForm


def voting_page(request):
    state = MachineState.get_instance()
    candidates = Candidate.objects.all()
    total_votes = Vote.objects.count()
    return render(request, 'voting/voting_page.html', {
        'candidates': candidates,
        'is_locked': state.is_locked,
        'total_votes': total_votes,
        'total_students': state.total_students,
    })


@csrf_exempt
@require_POST
def cast_vote(request):
    state = MachineState.get_instance()
    if state.is_locked:
        return JsonResponse({'success': False, 'message': 'Machine is locked. Wait for teacher.'}, status=400)

    candidate_id = request.POST.get('candidate_id')
    if not candidate_id:
        return JsonResponse({'success': False, 'message': 'Invalid request.'}, status=400)

    candidate = get_object_or_404(Candidate, pk=candidate_id)

    Vote.objects.create(candidate=candidate)
    Candidate.objects.filter(pk=candidate.id).update(vote_count=F('vote_count') + 1)

    state.is_locked = True
    state.save()

    return JsonResponse({
        'success': True,
        'message': 'Vote recorded!',
        'candidate_name': candidate.name,
    })


@csrf_exempt
@require_POST
def next_student(request):
    state = MachineState.get_instance()
    state.is_locked = False
    state.save()
    return JsonResponse({'success': True, 'message': 'Machine ready for next student.'})


@csrf_exempt
def machine_status(request):
    state = MachineState.get_instance()
    total_votes = Vote.objects.count()
    return JsonResponse({
        'is_locked': state.is_locked,
        'total_votes': total_votes,
    })


class AdminLoginView(View):
    def get(self, request):
        if request.user.is_authenticated:
            return redirect('voting:admin_dashboard')
        form = AdminLoginForm()
        return render(request, 'voting/admin_login.html', {'form': form})

    def post(self, request):
        form = AdminLoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('voting:admin_dashboard')
        return render(request, 'voting/admin_login.html', {'form': form})


@login_required
def admin_logout_view(request):
    logout(request)
    return redirect('voting:admin_login')


@login_required
def admin_dashboard(request):
    candidates = Candidate.objects.all().order_by('-vote_count')
    total_votes = Vote.objects.count()
    state = MachineState.get_instance()
    winner = candidates.first() if total_votes > 0 else None

    votes_by_date = (
        Vote.objects.annotate(date=TruncDate('timestamp'))
        .values('date')
        .annotate(count=Count('id'))
        .order_by('date')
    )

    recent_votes = Vote.objects.select_related('candidate').all()[:20]

    context = {
        'candidates': candidates,
        'total_votes': total_votes,
        'total_students': state.total_students,
        'winner': winner,
        'votes_by_date': list(votes_by_date),
        'recent_votes': recent_votes,
        'machine_state': state,
    }
    return render(request, 'voting/admin_dashboard.html', context)


@login_required
def result_page(request):
    candidates = Candidate.objects.all().order_by('-vote_count')
    total_votes = Vote.objects.count()
    winner = candidates.first() if total_votes > 0 else None

    candidate_data = []
    for c in candidates:
        pct = round((c.vote_count / total_votes * 100), 1) if total_votes > 0 else 0
        candidate_data.append({
            'candidate': c,
            'votes': c.vote_count,
            'percentage': pct,
        })

    context = {
        'candidate_data': candidate_data,
        'total_votes': total_votes,
        'winner': winner,
    }
    return render(request, 'voting/result_page.html', context)


@login_required
def export_pdf(request):
    try:
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Sumati Balwan School - Voting Report", styles['Title']))
        elements.append(Spacer(1, 0.3 * inch))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 0.2 * inch))

        total_votes = Vote.objects.count()
        data = [['Rank', 'Candidate', 'Votes', 'Percentage']]
        for i, c in enumerate(Candidate.objects.all().order_by('-vote_count'), 1):
            pct = f"{round(c.vote_count / total_votes * 100, 1)}%" if total_votes > 0 else "0%"
            data.append([str(i), c.name, str(c.vote_count), pct])

        table = Table(data, colWidths=[1*inch, 2*inch, 1.5*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#1a1a2e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), rl_colors.HexColor('#f8f9fa')),
            ('GRID', (0, 0), (-1, -1), 1, rl_colors.HexColor('#dee2e6')),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.3 * inch))
        elements.append(Paragraph(f"Total Votes: {total_votes}", styles['Normal']))

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="voting_report_{datetime.now().strftime("%Y%m%d")}.pdf"'
        return response

    except ImportError:
        return HttpResponse("PDF export requires reportlab. pip install reportlab", status=500)


@login_required
def export_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Voting Report"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")

        headers = ['Rank', 'Candidate', 'Votes', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        total_votes = Vote.objects.count()
        for row, c in enumerate(Candidate.objects.all().order_by('-vote_count'), 2):
            pct = round(c.vote_count / total_votes * 100, 1) if total_votes > 0 else 0
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=c.name)
            ws.cell(row=row, column=3, value=c.vote_count)
            ws.cell(row=row, column=4, value=f"{pct}%")

        for col in range(1, 5):
            ws.column_dimensions[chr(64 + col)].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="voting_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        return response

    except ImportError:
        return HttpResponse("Excel export requires openpyxl. pip install openpyxl", status=500)


@login_required
def print_result(request):
    candidates = Candidate.objects.all().order_by('-vote_count')
    total_votes = Vote.objects.count()
    winner = candidates.first() if total_votes > 0 else None

    candidate_data = []
    for c in candidates:
        pct = round((c.vote_count / total_votes * 100), 1) if total_votes > 0 else 0
        candidate_data.append({
            'candidate': c,
            'votes': c.vote_count,
            'percentage': pct,
        })

    context = {
        'candidate_data': candidate_data,
        'total_votes': total_votes,
        'winner': winner,
    }
    return render(request, 'voting/print_result.html', context)


@login_required
@require_POST
def reset_election(request):
    Vote.objects.all().delete()
    Candidate.objects.all().update(vote_count=0)
    state = MachineState.get_instance()
    state.is_locked = False
    state.save()
    messages.success(request, 'Election has been reset successfully.')
    return redirect('voting:admin_dashboard')


@login_required
def charts_data(request):
    candidates = Candidate.objects.all().order_by('-vote_count')
    total_votes = Vote.objects.count()
    labels = [c.name for c in candidates]
    votes_data = [c.vote_count for c in candidates]
    colors = [
        '#FFD700', '#8B4513', '#2196F3', '#FF9800', '#E91E63',
        '#4CAF50', '#F44336', '#009688', '#FF5722', '#00BCD4',
        '#3F51B5', '#9C27B0', '#FFC107', '#FF9800', '#607D8B',
        '#3F51B5', '#F44336', '#E91E63', '#03A9F4', '#FFA07A',
        '#FFD700', '#8B4513', '#2196F3', '#FF9800', '#E91E63',
        '#4CAF50', '#F44336', '#009688', '#FF5722', '#00BCD4',
        '#3F51B5', '#9C27B0', '#FFC107', '#FF9800', '#607D8B',
        '#3F51B5', '#F44336', '#E91E63', '#03A9F4', '#FFA07A',
    ]
    return JsonResponse({
        'labels': labels,
        'votes': votes_data,
        'colors': colors[:len(labels)],
        'total_votes': total_votes,
    })
